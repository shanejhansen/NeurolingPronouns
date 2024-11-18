import pandas as pd
import random
import logging
from typing import List, Tuple
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from collections import Counter

# Set up logging
logging.basicConfig(level=logging.DEBUG, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

character_counter = Counter()

def load_csv(filename: str) -> pd.DataFrame:
    """Load a CSV file and return a pandas DataFrame."""
    try:
        df = pd.read_csv(filename)
        logger.info(f"Successfully loaded {filename}")
        logger.debug(f"Columns in {filename}: {df.columns.tolist()}")
        logger.debug(f"First few rows of {filename}:\n{df.head().to_string()}")
        return df
    except Exception as e:
        logger.error(f"Error loading {filename}: {str(e)}")
        raise

def get_combination(item_num: int, characters: List[str], language: str) -> str:
    """Get a balanced combination of two characters for each item, with Alex first if included."""
    global character_counter
    
    if len(character_counter) == 0:  # Initialize counter if it's empty
        character_counter = Counter({char: 0 for char in characters})
    
    # Sort characters by usage count
    sorted_chars = sorted(characters, key=lambda x: (character_counter[x], random.random()))
    
    # Select the two least used characters
    selected = sorted_chars[:2]
    
    # Update counter
    character_counter.update(selected)
    
    # Ensure Alex is first if selected
    if "Alex" in selected:
        selected.remove("Alex")
        selected = ["Alex"] + selected
    
    return f"{selected[0]} {'et' if language == 'French' else 'and'} {selected[1]}"

def get_pronoun(condition: int, language: str) -> str:
    """Get the appropriate pronoun based on the condition and language."""
    pronouns = {
        "French": {1: "elle", 2: "il", 3: "il", 4: "elle", 5: "iel", 6: ["il", "elle"], 7: "iel", 8: ["il", "elle"]},
        "English": {9: "she", 10: "he", 11: "he", 12: "she", 13: "they", 14: ["he", "she"], 15: "they", 16: ["he", "she"]}
    }
    pronoun = pronouns[language][condition]
    if isinstance(pronoun, list):
        pronoun = random.choice(pronoun)
    return pronoun

def get_subject(condition: int) -> str:
    """Get the appropriate subject based on the condition."""
    subjects = {1: "Jennifer", 2: "Jennifer", 3: "George", 4: "George", 5: "Alex", 6: "Alex",
                9: "Jennifer", 10: "Jennifer", 11: "George", 12: "George", 13: "Alex", 14: "Alex"}
    return subjects.get(condition, None)

def get_french_translation(word: str, translations: pd.DataFrame) -> str:
    """Get the French translation of a word."""
    if word == "cooks":
        return "cuisine"
    if word in translations.index:
        return translations.loc[word, translations.columns[0]]
    elif word.endswith('s') and word[:-1] in translations.index:
        return translations.loc[word[:-1], translations.columns[0]]
    elif word.endswith('es') and word[:-2] in translations.index:
        return translations.loc[word[:-2], translations.columns[0]]
    logger.warning(f"No translation found for word: {word}")
    return word

def translate_and_conjugate_french(verb: str, translations: pd.DataFrame, plural: bool) -> str:
    """Translate an English verb to French and conjugate it."""
    logger.debug(f"Translating verb: {verb}")
    french_verb = get_french_translation(verb, translations)
    logger.debug(f"French translation: {french_verb}")
    if plural:
        if french_verb == "cuisine":
            return "cuisinent"
        elif french_verb.endswith('e'):
            return french_verb[:-1] + 'ent'
        elif french_verb.endswith('er'):
            return french_verb[:-2] + 'ent'
        else:
            return french_verb + 'ent'
    return french_verb

def get_base_form(verb: str) -> str:
    """Get the base form of a verb, considering it might be in third-person singular."""
    if verb.endswith('ies'):
        return verb[:-3] + 'y'
    elif verb.endswith('es'):
        return verb[:-2]
    elif verb.endswith('s'):
        return verb[:-1]
    return verb

def conjugate_english(verb: str, singular: bool) -> str:
    """Conjugate an English verb."""
    base_verb = get_base_form(verb)
    if not singular:
        return base_verb
    
    irregular_verbs = {
        "be": "is",
        "have": "has",
        "do": "does",
        "go": "goes",
        "cry": "cries"
    }
    if base_verb in irregular_verbs:
        return irregular_verbs[base_verb]
    elif base_verb.endswith('y'):
        return base_verb[:-1] + 'ies'
    elif base_verb.endswith(('s', 'ch', 'sh', 'x', 'z', 'o')):
        return base_verb + 'es'
    else:
        return base_verb + 's'

def correct_verb_spelling(verb: str) -> str:
    """Correct specific verb spellings after conjugation."""
    corrections = {
        "plaies": "plays",
        "wavs": "waves",
        "wav": "wave",
        "whistls": "whistles",
        "whistl": "whistle"
    }
    return corrections.get(verb, verb)

def generate_stimulus(item: pd.Series, condition: int, translations: pd.DataFrame) -> str:
    """Generate a stimulus based on the item, condition, and translations."""
    language = "French" if condition <= 8 else "English"
    characters = ["Jennifer", "George", "Alex"]
    
    is_plural = condition in [7, 8, 15, 16]
    subject = get_combination(item.name, characters, language) if is_plural else get_subject(condition)
    pronoun = get_pronoun(condition, language)
    
    try:
        if language == "French":
            verb1 = translate_and_conjugate_french(item['verb1'], translations, is_plural)
            verb2 = translate_and_conjugate_french(item['verb2'], translations, is_plural)
            connector = get_french_translation(item['connector'], translations)
        else:
            is_plural_or_they = is_plural or condition == 13
            verb1 = correct_verb_spelling(conjugate_english(item['verb1'], not is_plural))
            verb2 = correct_verb_spelling(get_base_form(item['verb2']) if is_plural_or_they else conjugate_english(item['verb2'], True))
            connector = item['connector']
        
        return f"{subject} {verb1} {connector} {pronoun} {verb2}."
    except Exception as e:
        logger.error(f"Error in generate_stimulus: {str(e)}")
        logger.debug(f"Item: {item.to_dict()}, Condition: {condition}")
        raise

def create_excel_file(stimuli: List[Tuple[int, int, str]], filename: str):
    """Create an Excel file with the stimuli."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Stimuli"

    # Add headers
    headers = ['Item Number', 'Condition Number', 'Stimulus']
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
        cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")

    # Add data
    for row, (item_num, condition, stimulus) in enumerate(stimuli, start=2):
        ws.cell(row=row, column=1, value=item_num)
        ws.cell(row=row, column=2, value=condition)
        ws.cell(row=row, column=3, value=stimulus)

    # Adjust column widths
    for column_cells in ws.columns:
        length = max(len(str(cell.value)) for cell in column_cells)
        ws.column_dimensions[column_cells[0].column_letter].width = length + 2

    # Save the workbook
    wb.save(filename)
    logger.info(f"Saved {filename}")

def main():
    try:
        # Load CSVs
        items = load_csv('items.csv')
        assignments = load_csv('item_assignment.csv')
        translations = load_csv('translations.csv')
        
        # Prepare translations DataFrame
        translations.set_index(translations.columns[0], inplace=True)  # Set English words as index
        logger.debug(f"Translations DataFrame index: {translations.index.tolist()[:5]}...")  # Show first 5 index items
        
        # Generate stimuli for each list
        for list_num in assignments.columns[1:]:
            logger.info(f"Generating stimuli for List {list_num}")
            stimuli = []
            
            # Reset character counter for each list
            global character_counter
            character_counter = Counter()
            
            for _, row in assignments.iterrows():
                try:
                    item_num = row['Item:']
                    condition = row[list_num]
                    
                    # Use the index to access the item, assuming item numbers start from 1
                    item = items.iloc[item_num - 1]
                    
                    stimulus = generate_stimulus(item, condition, translations)
                    stimuli.append((item_num, condition, stimulus))
                except Exception as e:
                    logger.error(f"Error generating stimulus for item {item_num} in List {list_num}: {str(e)}")
                    logger.debug(f"Item data: {item.to_dict() if 'item' in locals() else 'Item not found'}")
            
            # Create Excel file
            output_filename = f'stimuli_list_{list_num}.xlsx'
            create_excel_file(stimuli, output_filename)
    
    except Exception as e:
        logger.error(f"An error occurred: {str(e)}")
        logger.debug("DataFrame information:")
        logger.debug(f"Items DataFrame shape: {items.shape if 'items' in locals() else 'Not loaded'}")
        logger.debug(f"Assignments DataFrame shape: {assignments.shape if 'assignments' in locals() else 'Not loaded'}")
        logger.debug(f"Translations DataFrame shape: {translations.shape if 'translations' in locals() else 'Not loaded'}")

if __name__ == "__main__":
    main()